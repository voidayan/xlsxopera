from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict


class Notebook:

    def __init__(self, file: str, sheet: str = "active"):
        self.file = file
        self.sheet = sheet
        self.book = load_workbook(file).active \
            if sheet == "active" \
            else load_workbook(file)[sheet]

    def data_rows(self) -> List[str]:
        return [
            cell.value for row in self.book.iter_rows()
            for cell in row
        ]

    def data_cols(self) -> List[str]:
        return [
            cell.value for column in self.book.iter_cols()
            for cell in column
        ]

    def list_rows(self, start: int = 1, end: int = None) -> List[List]:
        return [
            [cell.value for cell in row]
            for row in self.book.iter_rows(min_row=start, max_row=end)
        ]

    def dict_rows(self, start: int = 1, end: int = None) -> Dict[int, List]:
        cells_value_dict = {}
        for index, row in enumerate(
                self.list_rows(start=start, end=end)
        ):
            cells_value_dict.setdefault(start + index, [])
            for value in row:
                cells_value_dict[start + index].append(value)
        return cells_value_dict

    def list_cols(self, start: int = 1, end: int = None) -> List[List]:
        return [
            [cell.value for cell in column]
            for column in self.book.iter_cols(min_row=start, max_row=end)
        ]

    def dict_cols(self, start: int = 1, end: int = None) -> Dict[int, List]:
        cells_value_dict = {}
        for index, col in enumerate(
                self.list_cols(start=start, end=end)
        ):
            cells_value_dict.setdefault(index + 1, [])
            for value in col:
                cells_value_dict[index + 1].append(value)
        return cells_value_dict

    def headers(self, row: int = 1) -> List[str]:
        return [
            cell.value for row in self.book.iter_rows(
                min_row=row, max_row=row
            ) for cell in row
        ]

    def value_rows(self, value: str) -> List[List[str]]:
        return [
            row for row in self.list_rows() if value in row
        ]

    def value_cols(self, value: str) -> List[List[str]]:
        return [
            col for col in self.list_cols() if value in col
        ]

    def data_convert(self, headers_row: int = 1) -> Dict[str, List]:
        convert_book = self.book
        headers = self.headers()
        cells_value_dict = {}
        for header in headers:
            if header is not None:
                cells_value_dict.setdefault(header, [])
            else:
                continue
            for row in [
                [cell.value for cell in row]
                for row in convert_book.iter_rows(
                    min_row=headers_row + 1
                )
            ]:
                cells_value_dict[header].append(
                    row[headers.index(header)]
                )
        return cells_value_dict

    def header_values(self, header: str) -> List[str]:
        header_values = []
        for index_row, header_row in enumerate(
                self.list_rows()
        ):
            if header in header_row:
                for index, cell_header in enumerate(header_row):
                    if cell_header == header:
                        for row in self.list_rows(
                                start=index_row + 2
                        ):
                            header_values.append(row[index])
        return header_values

    def find(self, value: str) -> bool:
        for cell_value in self.data_rows():
            if cell_value == value:
                return True
        return False

    def value_position(self, value: str = "") -> str:
        for row in self.book.iter_rows():
            for cell in row:
                if cell.value == value:
                    return f"{get_column_letter(cell.column)}{cell.row}"

    def value_coordinates(self, value: str = "") -> tuple:
        for row in self.book.iter_rows():
            for cell in row:
                if cell.value == value:
                    return cell.row, cell.column
